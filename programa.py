#!/usr/bin/env python3
from PIL import Image
from tkinter.ttk import Frame, Button
from tkinter import Tk, BOTH ,Listbox, StringVar, END , Menu , Label , Entry , DISABLED ,NORMAL, RAISED , FLAT , LEFT , TOP , X ,Canvas , Checkbutton , BooleanVar , IntVar , Radiobutton
from tkinter import messagebox as mbox
from datetime import datetime
from sqlalchemy import create_engine , DateTime , ForeignKey , Boolean , func
from sqlalchemy.orm import sessionmaker , relationship , backref
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column , Integer , Numeric , String
import time
import math as m
from sql.models import *

from tkinter import messagebox as mbox

import openpyxl
from openpyxl.styles import Font , Color
from correo.attached import sendmail
import os
from funciones.internet import is_connected

engine = create_engine('sqlite:///alimento_db')
Session = sessionmaker(bind=engine)

session = Session()

class Aplicacion(Frame):
    def __init__(self,parent):
        Frame.__init__(self , parent)
        
        self.parent = parent
        self.conexion = True#is_connected()
        self.initUI()

    def initUI(self):
        self.widgets_alta_usuario = []
        self.widgets_alimento = []
        self.widgets_orden = []
        self.widgets_ordenes_no_enviadas = []
        self.labels = []
        self.botones = []
        self.widgets_grafica = []
        self.widgets_estadistica = []
        self.diccionario = self.familias()
        self.meses = ["Ene" , "Feb" , "Mar" , "Abr" , "May" , "jun" , "Jul" , "Ago" , "Sep" , "Oct" , "Nov" , "Dic"]
        self.parent.title("Venta de Alimentos")
        for i in range(4):
            self.columnconfigure(i, pad=3)
        for i in range(4):
            self.rowconfigure(i, pad=3)
        self.toolbar()    
        self.menu()
        self.Orden()
        #self.listBox()
        sw = self.parent.winfo_screenwidth()
        sh = self.parent.winfo_screenheight()
        self.parent.geometry('%dx%d' % (sw/(1.7), sh/2))        
        self.pack()
    
    def redondear_operacion(self,magnitud,base = 0.5):
        return base * round(float(magnitud) / base)    

    def toolbar(self):
        toolbar = Frame(self.parent , relief = RAISED)
        eimg = Image.open('updater.png')
        Salir = Button(toolbar  ,text = "Salir" ,command = self.onExit )
        Salir.pack(side = LEFT )
        ordenes = Button(toolbar  ,text = "Ordenes" ,command = self.Orden )
        ordenes.pack(side = LEFT )
        usuarios = Button(toolbar  , text = "Usuarios" ,command = self.alta_usuario)
        usuarios.pack(side = LEFT )
        alimentos = Button(toolbar  , text = "Alimentos", command = self.Alimento)
        alimentos.pack(side = LEFT )
        pendientes = Button(toolbar  , text = "Pendientes", command = self.Ordenes_no_enviadas)
        pendientes.pack(side = LEFT )
        graficas = Button(toolbar  , text = "Graficas", command = lambda mes = datetime.now().month , anio = datetime.now().year: self.graficas(mes , anio))
        graficas.pack(side = LEFT )
        estadisticas = Button(toolbar  , text = "Estadisticas", command = self.estadisticas)
        estadisticas.pack(side = LEFT )
        toolbar.pack(side = TOP , fill = X)




    def menu(self):
        menubar = Menu(self.parent)
        self.parent.config(menu = menubar)

        filemenu = Menu(menubar)

        submenu = Menu(filemenu)
        submenu.add_command(label = "Desde archivo " , command = self.onExit)
        submenu.add_command(label = "Modificar Usuario")
        submenu.add_command(label = "Borrar Usuario")

        filemenu.add_cascade(label = "Importar" , menu = submenu , underline = 0)

        submenu = Menu(filemenu)
        submenu.add_command(label = "A archivo Excel... " , command = self.onExit)
        submenu.add_command(label = "A archivo de texto...")
        submenu.add_command(label = "Mandar por email...")

        filemenu.add_cascade(label = "Exportar" , menu = submenu , underline = 0)

        filemenu.add_separator()

        filemenu.add_command(label = "Salir" ,underline=0, command = self.onExit)
        menubar.add_cascade(label = "Archivo" ,underline=0 ,menu = filemenu)

        '''EDITAR '''

        filemenu = Menu(menubar)
        
        submenu = Menu(filemenu)
        submenu.add_command(label = "Nuevo usuario " , command = self.alta_usuario)
        submenu.add_command(label = "Modificar Usuario")
        submenu.add_command(label = "Borrar Usuario")

        filemenu.add_cascade(label = "Usuario" , menu = submenu , underline = 0)

        submenu = Menu(filemenu)
        submenu.add_command(label = "Nueva orden " , command = self.Orden)
        submenu.add_command(label = "Modificar Orden")
        submenu.add_command(label = "Borrar Orden")

        filemenu.add_cascade(label = "Orden" , menu = submenu , underline = 0)

        submenu = Menu(filemenu)
        submenu.add_command(label = "Nuevo Producto " , command = self.Alimento)
        submenu.add_command(label = "Modificar Producto")
        submenu.add_command(label = "Borrar Producto")

        filemenu.add_cascade(label = "Producto" , menu = submenu , underline = 0)

        filemenu.add_separator()

        filemenu.add_command(label = "Exit" ,underline=0, command = self.onExit)
        menubar.add_cascade(label = "Editar" ,underline=0 ,menu = filemenu)

        '''CONTROL'''

        filemenu = Menu(menubar)
        
        submenu = Menu(filemenu)
        submenu.add_command(label = "Reporte..." , command = self.onExit)
        submenu.add_command(label = "Escribir a archivo...")
        submenu.add_command(label = "Enviar por Email...")
        submenu.add_command(label = "Reporte interactivo...")


        filemenu.add_cascade(label = "Inventario" , menu = submenu , underline = 0)

        
        menubar.add_cascade(label = "Control" ,underline=0 ,menu = filemenu)

        '''Estadistico '''

        filemenu = Menu(menubar)
        
        submenu = Menu(filemenu)
        submenu.add_command(label = "Nuevo usuario " , command = self.onExit)
        submenu.add_command(label = "Modificar Usuario")
        submenu.add_command(label = "Borrar Usuario")

        filemenu.add_cascade(label = "Usuario" , menu = submenu , underline = 0)

        submenu = Menu(filemenu)
        submenu.add_command(label = "Nueva orden " , command = self.onExit)
        submenu.add_command(label = "Modificar Orden")
        submenu.add_command(label = "Borrar Orden")

        filemenu.add_cascade(label = "Orden" , menu = submenu , underline = 0)

        submenu = Menu(filemenu)
        submenu.add_command(label = "Nuevo Producto " , command = self.onExit)
        submenu.add_command(label = "Modificar Producto")
        submenu.add_command(label = "Borrar Producto")

        filemenu.add_cascade(label = "Producto" , menu = submenu , underline = 0)

        filemenu.add_separator()

        filemenu.add_command(label = "Exit" ,underline=0, command = self.onExit)
        menubar.add_cascade(label = "Estadistico" ,underline=0 ,menu = filemenu)

        '''CONFIGURACION'''

        filemenu = Menu(menubar)
        
        


        filemenu.add_command(label = "Configuracion general" )

        
        menubar.add_cascade(label = "Configuracion" ,underline=0 ,menu = filemenu)

    
    def reporte_ventas(self):

        respuesta = mbox.askquestion("Generar Excel", "Deseas Generar reporte de ventas a un archivo xlsx?")
        if respuesta == 'yes':
            clr = Color(rgb = '00204040')
            fontTitulo = Font(size = 24 , bold = True , color = clr)
            fontEncabezado = Font(bold = True)

            fecha = datetime.now()
            wb = openpyxl.Workbook()
            sheet = wb.active
            wb.create_sheet(index=0, title='Ventas de Hoy')
            wb.create_sheet(index=1, title='Ventas Este mes')
            wb.create_sheet(index=2, title='Ventas Anuales')

    
            for nombre in wb.get_sheet_names():
                sheet = wb.get_sheet_by_name(nombre)
                #sheet.title = 'Venta %s-%s-%s'%(fecha.day,self.meses[fecha.month],fecha.year)
                sheet['C1'].value = nombre.upper()
                sheet['C1'].font = fontTitulo
                sheet.merge_cells('C1:F1')
                encabezado = ['Usuario','Orden','Linea','Articulo','Precio unitario','Total','Ganancia','Fecha']
                sheet['B3'].value = 'Venta'
                sheet['B3'].font = fontEncabezado
                sheet['E3'].value = 'Ganancia'
                sheet['E3'].font = fontEncabezado
                LINEA_ANCHA = ['A','D','E','H']
                for columna in LINEA_ANCHA:
                    sheet.column_dimensions[columna].width = 20


        

                for column in range(len(encabezado)):
                    sheet.cell(row = 6 , column = column+1).value = encabezado[column]
                    sheet.cell(row = 6 , column = column+1).font = fontEncabezado

                sheet.freeze_panes = 'A7'
            for hoja in range(len(wb.get_sheet_names())):
                sheet = wb.get_sheet_by_name(wb.get_sheet_names()[hoja])

                usuarios = session.query(Usuario).all()
                fecha_act = datetime.now()
                row = 7
            
                for usuario in usuarios:
                    for orden in usuario.ordenes:
                        fecha = datetime.date(orden.fecha)
                        if sheet == wb.get_sheet_by_name(wb.get_sheet_names()[0]):
                            if fecha.day == fecha_act.day and fecha.month == fecha_act.month and fecha.year == fecha_act.year and orden.enviado:
                                for linea in orden.lineas:
                                    print("Diario")
                                    sheet.cell(row = row , column = 1).value = usuario.nombre
                                    sheet.cell(row = row , column = 2).value = orden.id
                                    sheet.cell(row = row , column = 3).value = linea.cantidad
                                    sheet.cell(row = row , column = 4).value = linea.alimento.nombre
                                    sheet.cell(row = row , column = 5).value = (linea.costo_extendido+linea.alimento.costo)
                                    sheet.cell(row = row , column = 6).value = (linea.costo_extendido+linea.alimento.costo)*linea.cantidad
                                    sheet.cell(row = row , column = 7).value = linea.costo_extendido*linea.cantidad
                                    sheet.cell(row = row , column = 8).value = orden.fecha
                                    row += 1
    
                        if sheet == wb.get_sheet_by_name(wb.get_sheet_names()[1]):
                            if fecha.month == fecha_act.month and fecha.year == fecha_act.year and orden.enviado:
                                for linea in orden.lineas:
                                    print("MES")
                                    sheet.cell(row = row , column = 1).value = usuario.nombre
                                    sheet.cell(row = row , column = 2).value = orden.id
                                    sheet.cell(row = row , column = 3).value = linea.cantidad
                                    sheet.cell(row = row , column = 4).value = linea.alimento.nombre
                                    sheet.cell(row = row , column = 5).value = (linea.costo_extendido+linea.alimento.costo)
                                    sheet.cell(row = row , column = 6).value = (linea.costo_extendido+linea.alimento.costo)*linea.cantidad
                                    sheet.cell(row = row , column = 7).value = linea.costo_extendido*linea.cantidad
                                    sheet.cell(row = row , column = 8).value = orden.fecha
                                    row += 1
                        if sheet == wb.get_sheet_by_name(wb.get_sheet_names()[2]):
                            if fecha.year == fecha_act.year and orden.enviado:
                                for linea in orden.lineas:
                                    print('ANUAL')
                                    sheet.cell(row = row , column = 1).value = usuario.nombre
                                    sheet.cell(row = row , column = 2).value = orden.id
                                    sheet.cell(row = row , column = 3).value = linea.cantidad
                                    sheet.cell(row = row , column = 4).value = linea.alimento.nombre
                                    sheet.cell(row = row , column = 5).value = (linea.costo_extendido+linea.alimento.costo)
                                    sheet.cell(row = row , column = 6).value = (linea.costo_extendido+linea.alimento.costo)*linea.cantidad
                                    sheet.cell(row = row , column = 7).value = linea.costo_extendido*linea.cantidad
                                    sheet.cell(row = row , column = 8).value = orden.fecha
                                    row += 1   
            for hoja in wb.get_sheet_names():
                sheet = wb.get_sheet_by_name(hoja)
                print(sheet.max_row)
                sheet['F3'].value = '=SUM(G7:G%s)'%sheet.max_row
                sheet['C3'].value = '=SUM(F7:F%s)'%sheet.max_row      
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))                                                                                   
            wb.save('reporte_ventas.xlsx')    
              
            if os.path.isfile(os.getcwd()+'/reporte_ventas.xlsx'):
                fecha = datetime.now()
                sendmail('Pocoyo',['luis@4suredesign.com','lic.myriamdelgado@gmail.com'],
                	     'Reporte de ventas %s-%s-%s'%(fecha.day,self.meses[fecha.month-1],fecha.year),
                	     'Correo generado desde sistema de ventas ',
                	     files = [os.getcwd()+'/reporte_ventas.xlsx'])
                mbox.showinfo('Archivo generado ' , 'Se genero exitosamente el archivo y se envio por email...')
            else: print(os.getcwd()+'/reporte_ventas.xlsx'+ 'No es un archivo valido')          
            

    def estadisticas(self):
        print("estadistica...")
        self.destruir()
        self.busqueda_usuarios = Entry(self ,)
        self.busqueda_usuarios.grid(row = 1 , column = 1 , sticky = 'NSEW')
        self.busqueda_usuarios.bind('<Return>' , self.filtro_busqueda_usuarios)
        self.variable = IntVar()

        MODOS = [
                 ("Hoy",1),
                 ("Este Mes",2),
                 ("Este Anio",3),]
        self.variable.set(2)
        self.botones = []
        for texto, valor in MODOS:
            self.boton = Radiobutton(self , text = texto , variable = self.variable , value = valor , command = self.prueba_joins)
            self.boton.grid(column = 0)         
            self.botones.append(self.boton)

        self.usuarios_lb = Listbox(self ,height = 10 , font = (None , 15))
        self.filtro_busqueda_usuarios()
        self.usuarios_lb.bind("<<ListboxSelect>>", self.prueba_joins)
        self.usuarios_lb.grid(row = 2 , column = 1 , rowspan = 10)

        self.mayores_lb = Listbox(self ,height = 10 , font = (None , 15))
        self.mayores_lb.grid(row = 2 , column = 2 , rowspan = 10)

        boton = Button(self, text = "Reporte de Ventas en excel" , command = self.reporte_ventas)
        boton.grid(row = 20 , column = 2)
        self.widgets_estadistica = [self.busqueda_usuarios  , self.usuarios_lb , self.mayores_lb , boton]
        

    def poblar_maximos(self,matriz):
        
        self.mayores_lb.delete(0,END)
        for valor in matriz:
            self.mayores_lb.insert(END,valor)
        for i in range(0,self.mayores_lb.size(),2):
            self.mayores_lb.itemconfigure(i,background = '#f0f0ff')
        self.mayores_lb.grid()

    def desplegar_producto_mas_comprado_por_usuario(self , val):
        sender = val.widget
        idx = sender.curselection()
        nombre = sender.get(idx)
        usuario = session.query(Usuario).filter(Usuario.nombre == nombre).one()
        print(usuario.nombre)
        art_usuario = self.get_elemento_mas_comprado(user = usuario)
        art_mes = self.get_elemento_mas_comprado(mes = datetime.now().month)
        art_anio = self.get_elemento_mas_comprado(anio = datetime.now().year)

        self.user_est.set("""Usuario : {}\
        	\nArticulo mas Comprado \
        	\nEste mes : {:6} - {}\
        	\nEste anio :{:6} - {}""".format(usuario.nombre,
        		                             art_usuario['cantidad'] , session.query(Alimento).filter(Alimento.id == art_usuario['id']).one().nombre,
        		                             art_anio['cantidad'] , session.query(Alimento).filter(Alimento.id == art_anio['id']).one().nombre,
        		                             ))

        
    def prueba_joins(self , val = False):
        if not val: nombre = self.usuarios_lb.get(self.usuarios_lb.curselection())
        else:
            sender = val.widget
            idx = sender.curselection()
            nombre = sender.get(idx)
        query = session.query(Usuario.nombre ,Orden.id ,Alimento.nombre , Linea.cantidad ,Orden.fecha)
        query = query.join(Orden).join(Linea).join(Alimento)
        resultado = query.filter(Usuario.nombre == nombre).all()
        diccionario = {}
        for nombre,orden_id,alimento_nombre,cantidad,fecha in resultado:
            f = datetime.date(fecha)
            print("{:5} - {} - {} --{}".format(cantidad , alimento_nombre , nombre,f.month))
            if self.variable.get() == 1:
                print(1)
                if datetime.now().day == f.day:
                    if not alimento_nombre in diccionario.keys():
                        diccionario[alimento_nombre] = cantidad
                    else:
                        diccionario[alimento_nombre] += cantidad
            elif self.variable.get() == 2:
                if datetime.now().month == f.month:
                    if not alimento_nombre in diccionario.keys():
                        diccionario[alimento_nombre] = cantidad
                    else:
                        diccionario[alimento_nombre] += cantidad

            elif self.variable.get() == 3:
                if datetime.now().year == f.year:
                    if not alimento_nombre in diccionario.keys():
                        diccionario[alimento_nombre] = cantidad
                    else:
                        diccionario[alimento_nombre] += cantidad                        
            

        #print(diccionario)
        matriz = []
        for i in range(len(diccionario)):
            print(diccionario)
            max_key = max(diccionario,key=lambda x:diccionario[x])
            matriz.append("{:4} - {}".format(diccionario[max_key],max_key))

            diccionario.pop(max_key)

        
        print(matriz)    
        self.poblar_maximos(matriz[:5])    




    def get_elemento_mas_comprado(self , user = None , anio = None , mes = None):
        
        diccionario = {}
        ordenes = []
        
        if user:
            for orden in user.ordenes:
                if orden.enviado:

                    ordenes.append(orden)
            #ordenes = user.ordenes
        else:
            ordenes = session.query(Orden).filter(Orden.enviado == True)
            if anio and not mes:
                ordenes2 = []
                for orden in ordenes:
                    if orden.fecha.year == anio:
                        ordenes2.append(orden)
                ordenes = ordenes2
            if mes and not anio:
                ordenes2 = []
                for orden in ordenes:
                    if orden.fecha.month == mes:
                        ordenes2.append(orden)
                ordenes = ordenes2
                    	
        for orden in ordenes:
            for linea in orden.lineas:

                if linea.alimento.id in diccionario.keys():
                    diccionario[linea.alimento.id] += linea.cantidad
                else :
                    diccionario[linea.alimento.id] = linea.cantidad
        print(diccionario)
        print(max(diccionario.keys()))
        maximo = 0

        for id_ , valor in diccionario.items():
            print("{:3} - {}".format(id_,valor))            
            if valor > maximo: 
                maximo = valor
                id = id_
        
        return dict(cantidad = maximo , id = id)


    def graficas(self , mes , anio):
        print("Variables de entrada | mes = %s anio = %s"%(mes , anio))
        self.destruir()
        ordenes_enviadas = session.query(Orden).filter(Orden.enviado == True).all()
        costos = []
        ganancias = []
        venta_neta = []
        labels = []
        ganancia_diaria = {}
        self.checkbuttons = []
        for orden in ordenes_enviadas:
            if orden.fecha.month == mes and orden.fecha.year == anio:
                for linea in orden.lineas:
                    if orden.fecha.day in ganancia_diaria.keys():
                        ganancia_diaria[orden.fecha.day] += float(linea.costo_extendido)*float(linea.cantidad)
                    else:
                        ganancia_diaria[orden.fecha.day] = float(linea.costo_extendido)*float(linea.cantidad)  

        self.canvas = Canvas(self , width = 900 , height = 300 ,bg = 'black')
        self.canvas.grid(row = 1 , column = 1 ,columnspan = 31 , rowspan = 13)
        #labels = []
        column = 1
        print('esteeeeeeeeee')
        #for dia in range(31):
        #    labels.append(Label(self , text = str(dia+1)))
        #    labels[-1].grid(row = 14 , column = column)
        #    column += 1
        self.boolean_var = []
        mes_id = 1
        for mes in self.meses:
            #self.boolean_var.append(BooleanVar)
            self.checkbuttons.append(Button(self , text = mes+' '+str(anio) , command = lambda mes = mes_id , anio = anio :self.graficas(mes , anio) ))    
            #self.checkbuttons[-1].select()
            self.checkbuttons[-1].grid(column = 33 , row = mes_id)
            mes_id += 1
            #print(self.boolean_var)                  
                
        #ganancias = [round((x*div_maxi)+1) for x in range(13)]   
        print(ganancia_diaria)
        if not ganancia_diaria: 
            print(anio)
            mbox.showwarning("Busqueda", "No existen datos en la fecha especificada .")
            self.canvas.delete()
            #self.graficas(mes , anio)
        else:
            maximo = max(ganancia_diaria.values())
            dia_pasado = 0
            ganancia_pasada = 0
            
            for dia in range(1,32):
                if dia not in ganancia_diaria.keys():
                    self.canvas.create_line(dia_pasado*(900.0/31.0),300-(ganancia_pasada*300)/(maximo) , dia*(900.0/31) , 300-0,fill = 'red')
                    #print((dia_pasado,ganancia_pasada , dia , 300))
                    ganancia_pasada = 0
                else:
                    self.canvas.create_line(dia_pasado*(900.0/31.0),300-(ganancia_pasada*300)/(maximo) , dia*(900.0/31.0) , 300-(ganancia_diaria[dia]*300)/(maximo) , fill = 'red')
                    #print((dia_pasado,ganancia_pasada , dia , ganancia_diaria[dia]))
                    ganancia_pasada = ganancia_diaria[dia]
                dia_pasado = dia
                
        
    
            print('Entro a graficas')
    
            #self.canvas.grid(row = 1 , column = 1 ,columnspan = 31 , rowspan = 13)
            self.labels = []
            column = 1

            #for dia in range(31):
            #    self.labels.append(Label(self , text = str(dia+1)))
            #    self.labels[-1].grid(row = 14 , column = column)
            #    column += 1
            self.boolean_var = []
            mes_id = 1
            for mes in self.meses:
                #self.boolean_var.append(BooleanVar)
                self.checkbuttons.append(Button(self , text = mes+' '+str(anio) , command = lambda mes = mes_id , anio = anio :self.graficas(mes , anio) ))    
                #self.checkbuttons[-1].select()
                self.checkbuttons[-1].grid(column = 33 , row = mes_id)
                mes_id += 1
                #print(self.boolean_var)

            self.colors = ['red' , 'blue' , 'green' , 'yellow' , 'white'  , 'magenta' , 'orange' , 'pink' , 'brown' , 'gold' , 'silver' , '#f0f0ff']
            self.widgets_grafica = [self.canvas]

        



    def buscar_ordenes(self , clave):
        clave = self.busqueda.get().lower()
        usuarios = session.query(Usuario).filter(Usuario.nombre.like('%'+clave+'%') )        
        #for orden in usuarios.ordenes:
        #    print(orden.usuario.nombre)
        self.ordenes_lb.delete(0,END)
        
        for usuario in usuarios:
            for orden in usuario.ordenes:
                print(orden.enviado)
                if not orden.enviado:
                    fecha = "%s-%s-%s"%(orden.fecha.day,
            	 	                    orden.fecha.month,
            		                    orden.fecha.year,
            		                    )
                    self.ordenes_lb.insert(END,"{:10} - {}".format(fecha , orden.usuario.nombre+','+str(orden.id)))
        for i in range(0,self.ordenes_lb.size(),2):
            self.ordenes_lb.itemconfigure(i,background = '#f0f0ff')

          

    def poblar_ordenes(self):
        self.ordenes_lb.delete(0,END)
        for orden in session.query(Orden):
            if not orden.enviado:
                fecha = "%s-%s-%s"%(orden.fecha.day,
            		                orden.fecha.month,
            		                orden.fecha.year,
            		                )
                self.ordenes_lb.insert(END,"{:10} - {}".format(fecha , orden.usuario.nombre+','+str(orden.id)))
                
       
            
        for i in range(0,self.ordenes_lb.size(),2):
            self.ordenes_lb.itemconfigure(i,background = '#f0f0ff')
        
    def onSelectOrdenes(self ,event ):
        seleccion = self.ordenes_lb.get(self.ordenes_lb.curselection()).split(',')
        #print(seleccion)
        #print("id = %s nombre = %s"%(seleccion[1],seleccion[0][13:]))
        orden_id = seleccion[1]
        nombre = seleccion[0][13:]
        string = ''
        orden = session.query(Orden).filter(Orden.id == int(orden_id)).one()
        self.articulos_lb.delete(0,END)
        for linea in orden.lineas:
            self.articulos_lb.insert(END,"{:3}  {}".format(linea.cantidad , linea.alimento.nombre))
            
        for i in range(0,self.articulos_lb.size(),2):
            self.articulos_lb.itemconfigure(i,background = '#f0ffff')
        
    	


        
    def Ordenes_no_enviadas(self):
        self.destruir()

        titulo = Label(self , text = 'Consulta de Ordenes Pendientes' , font = (None , 20))
        titulo.grid(column = 0 , row = 1 , columnspan = 2)
        
        
        articulos = ''
        self.busqueda = Entry(self)
        self.busqueda.grid(row = 2 , column = 0 , columnspan = 2 , sticky = 'NSEW')
        self.ordenes_lb = Listbox(self , height = 10 , font = (None , 12))
        self.ordenes_lb.bind("<<ListboxSelect>>", self.onSelectOrdenes)
        self.poblar_ordenes()
        self.ordenes_lb.grid(column = 0 , row = 3 , columnspan = 2 , sticky = 'NSEW')
        self.string_articulos = StringVar()
        self.articulos_lb = Listbox(self , height = 10 , font = (None , 12))
        #self.string_articulos.set('Hola')
        self.articulos_lb.grid(column = 2 , row = 3 , columnspan = 2 , sticky = 'NSEW')
        self.boton = Button(self , text = 'Marcar como enviado' , command = self.marcar_enviado)
        self.boton.grid(column = 2 , row = 4 )
        self.busqueda.bind('<Return>' , self.buscar_ordenes)
        self.widgets_ordenes_no_enviadas = [titulo , self.busqueda , self.ordenes_lb ,
                                            self.articulos_lb , self.boton]

    def marcar_enviado(self ):
        seleccion = self.ordenes_lb.get(self.ordenes_lb.curselection()).split(',')
        orden_id = seleccion[1]
        nombre = seleccion[0][13:]

        orden = session.query(Orden).filter(Orden.id == orden_id).one()
        orden.enviado = True
        session.add(orden)
        session.commit()
        self.Ordenes_no_enviadas()
        

                

    def Orden(self):
        self.destruir()
        self.lineas = []
        lbl_titulo = Label(self , text = "          Alta de Ordenes            " ,font = (None , 30) )
        lbl_titulo.grid(row = 0 , column = 2 , columnspan = 3 )
        lbl_usuario= Label(self , text = 'Usuario' ,font = (None , 12))
        lbl_usuario.grid(row = 1 , column = 4 , sticky = "E")

        self.usuario = Entry(self)
        self.usuario.grid(row = 1 , column = 5)

        self.buscar_usuario = Button(self , text = 'Buscar Usuario' ,command = self.tk_busqueda_usuario)
        self.buscar_usuario.grid(row = 1 , column = 6 , sticky = 'NS')

        self.buscar_articulo = Button(self , text = 'Buscar Articulo' , command = self.tk_busqueda_producto)
        self.buscar_articulo.grid(row = 2 , column = 6 , rowspan = 2)

        lbl_familia= Label(self , text = 'Familia' ,font = (None , 12))
        lbl_familia.grid(row = 2 , column = 4 , sticky = "E")
        lbl_numero = Label(self , text = 'Numero' , font = (None , 12))
        lbl_numero.grid(row = 3 , column = 4 , sticky = "E")

        self.string_articulo = StringVar()
        self.lbl_articulo = Label(self , text = "hsdjkfhdskjhf" ,textvariable = self.string_articulo,font = (None , 12))
        self.lbl_articulo.grid(row = 5 , column = 2 , columnspan = 6 , sticky = 'WESN')

        self.familia = Entry(self)
        self.familia.grid(row = 2 , column = 5)
        self.numero = Entry(self)
        self.numero.grid(row = 3 , column = 5)
        self.lbl_cantidad = Label(self , text = 'Cantidad')
        self.lbl_cantidad.grid(row = 5 , column = 1)
        self.cantidad = Entry(self )
        self.cantidad.grid(row = 5 , column = 2 , sticky = 'W')
        self.cantidad.config(state = DISABLED)

        self.articulos_lb = Listbox(self , height = 10 , font = (None , 15))
        self.articulos_lb.bind("<<ListboxSelect>>", self.onSelectArticulos)
        self.articulos_lb.grid(row = 6 , column = 1 , columnspan = 6 , sticky = 'NSEW')
        
        self.string_subtotal = StringVar()
        self.lbl_subtotal = Label(self , text = '{:8} :'.format('Subtotal') , textvariable = self.string_subtotal,font = (None , 12))
        self.lbl_subtotal.grid(row = 7 ,column = 5)
        self.string_subtotal.set("{:8}:".format("Subtotal"))

        self.string_iva = StringVar()
        self.lbl_iva = Label(self , text = '{:8} :'.format('Iva') ,textvariable = self.string_iva ,font = (None , 12))
        self.lbl_iva.grid(row = 8 ,column = 5)  
        self.string_iva.set("{:8}:".format("IVA"))      
        self.string_total = StringVar()
        self.lbl_total = Label(self , text = '{:8} :'.format('Total') ,textvariable = self.string_total, font = (None , 12))
        self.lbl_total.grid(row = 9 ,column = 5)
        self.string_total.set("{:8}:".format("Total"))
        self.usuario.bind('<Return>',self.focusFamilia)
        self.familia.bind('<Return>',self.focusNumero)
        #self.familia.bind('<Enter>',self.focusNumero)
        self.numero.bind('<Return>' , self.desplegarArticulo)
        #self.numero.bind('<Enter>' , self.desplegarArticulo)
        self.cantidad.bind('<Return>',self.anadir_a_orden)
        self.parent.bind('<Control-c>' , self.finalizar_pedido)

        self.widgets_orden = [lbl_titulo , lbl_usuario , self.usuario ,self.buscar_usuario ,
                              self.buscar_articulo , lbl_familia , lbl_numero ,
                              self.lbl_articulo , self.familia , self.numero ,
                              self.lbl_cantidad ,self.cantidad, self.articulos_lb ,self.lbl_subtotal ,
                              self.lbl_iva , self.lbl_total]

    def tk_busqueda_usuario(self):
        self.ventana = Tk()
        self.ventana.title('Busqueda de usuario')
        #self.ventana.geometry('300x300')
        self.usuarios_lb = Listbox(self.ventana , height = 10 , font = (None , 15))
        self.usuarios_lb.bind("<<ListboxSelect>>", self.onSelect_busqueda_usuario)

        
        self.busqueda_usuarios = Entry(self.ventana)
        self.busqueda_usuarios.grid(columnspan = 2 , sticky = 'NSEW')
        self.busqueda_usuarios.bind('<Return>' , self.filtro_busqueda_usuarios)
        usuarios = session.query(Usuario).all()
        self.usuarios_lb.delete(0,END)
        for usuario in usuarios:
            self.usuarios_lb.insert(END,usuario.nombre)
        for i in range(0,self.usuarios_lb.size(),2):
            self.usuarios_lb.itemconfigure(i,background = '#f0f0ff')
        self.usuarios_lb.grid(sticky = 'NSEW')
    def filtro_busqueda_usuarios(self , *args):
        clave = self.busqueda_usuarios.get().lower()
        usuarios = session.query(Usuario).filter(Usuario.nombre.like('%'+clave+'%'))
        self.usuarios_lb.delete(0,END)
        for usuario in usuarios:
            self.usuarios_lb.insert(END,usuario.nombre)
        for i in range(0,self.usuarios_lb.size(),2):
            self.usuarios_lb.itemconfigure(i,background = '#f0f0ff')
        self.usuarios_lb.grid()    
    def onSelect_busqueda_usuario(self , val):
        sender = val.widget
        idx = sender.curselection()
        nombre = sender.get(idx)
        usuario = session.query(Usuario).filter(Usuario.nombre == nombre).one()
        
        self.usuario.delete(0,END)
        self.usuario.insert(0,usuario.id)
        self.focusFamilia()
        try:

            self.ventana.destroy()
        except:
            pass
        
        
    def tk_busqueda_producto(self):
        self.ventana = Tk()
        self.ventana.title('Busqueda de Alimento')
        #self.ventana.geometry('300x300')                
        self.alimentos_lb = Listbox(self.ventana ,height = 10 , font = (None , 15) )
        self.alimentos_lb.bind("<<ListboxSelect>>", self.onSelect_busqueda_alimento)

        
        self.busqueda_alimentos = Entry(self.ventana)
        self.busqueda_alimentos.grid(row = 0 , column = 0 , columnspan = 2 , sticky = 'NWES')
        self.busqueda_alimentos.bind('<Return>' , self.filtro_busqueda_alimentos)
        alimentos = session.query(Alimento).all()
        self.alimentos_lb.delete(0,END)
        for alimento in alimentos:
            self.alimentos_lb.insert(END,alimento.nombre)
        for i in range(0,self.alimentos_lb.size(),2):
            self.alimentos_lb.itemconfigure(i,background = '#f0f0ff')
        self.alimentos_lb.grid(sticky = 'NSEW' , row = 1 , column = 0 ,columnspan = 2 )          

    def filtro_busqueda_alimentos(self , *args):
        clave = self.busqueda_alimentos.get().lower()
        alimentos = session.query(Alimento).filter(Alimento.nombre.like('%'+clave+'%'))
        self.alimentos_lb.delete(0,END)
        for alimento in alimentos:
            self.alimentos_lb.insert(END,alimento.nombre)
        for i in range(0,self.alimentos_lb.size(),2):
            self.alimentos_lb.itemconfigure(i,background = '#f0f0ff')
        self.alimentos_lb.grid() 

    def onSelect_busqueda_alimento(self , val):
        sender = val.widget
        idx = sender.curselection()
        nombre = sender.get(idx)
        alimento = session.query(Alimento).filter(Alimento.nombre == nombre).one()
        familia = int(int(alimento.sku)/1000)
        numero = int(alimento.sku)%1000



        self.familia.delete(0,END)
        self.familia.insert(0,familia)
        self.numero.delete(0,END)
        self.numero.insert(0,numero)
        self.focusNumero()
        try:

            self.ventana.destroy()
        except:
            pass        



    def finalizar_pedido(self , *args):
        if not self.lineas:
            mbox.showwarning("Busqueda", "No existen Articulos en linea \n favor de agregar articulos .")
            self.focusFamilia()
        else :
            usuario_id = self.usuario.get()
            usuario = session.query(Usuario).filter(Usuario.id == usuario_id).one()
            orden = Orden(lineas = self.lineas)
            orden.fecha = datetime.now()
            if usuario.id == 1 : orden.enviado = True
            
            usuario.ordenes.append(orden)
            print('alimento desde usuario : \n>>%s'%usuario.ordenes[0].lineas[0].alimento.nombre)
            session.add(usuario)
            session.commit()
            mbox.showinfo("Alta de Orden Exitosa", "Se agrego la orden a la base de datos , \n tambien se envio por correo")
            self.articulos_lb.delete(0,END)
            if self.conexion:
                for linea in orden.lineas:
                    print('{:4} - {}'.format(linea.cantidad,linea.alimento.nombre))

            


        
  

    def anadir_a_orden(self , *args):
        familia = self.familia.get()
        numero = self.numero.get()

        cantidad = self.cantidad.get()
        ent_sku = int(int(familia)*1000 + int(numero))
        self.cantidad.delete(0,END)
        self.cantidad.config(state = DISABLED)
        self.familia.delete(0,END)
        self.numero.delete(0,END)

        alimento = session.query(Alimento).filter(Alimento.sku == ent_sku).one()
        self.lineas.append(Linea())
        self.lineas[-1].alimento = alimento




        if not cantidad : cantidad = 1
        cantidad = int(float(cantidad))
        
        #print(cantidad)
        usuario = self.usuario.get()

        #print("Usuario = %s"%self.usuario.get())
        if int(cantidad) < 1:
            cantidad = 1
        
        self.lineas[-1].cantidad = cantidad
        costo_alimento = self.lineas[-1].alimento.costo
        #print(costo_alimento)
        self.lineas[-1].costo_extendido = float(costo_alimento)*0.3
        #print(self.lineas[-1].costo_extendido)

        alimento_nombre = self.lineas[-1].alimento.nombre
        self.articulos_lb.insert(END,"{:3} - {} -${}".format(cantidad ,alimento_nombre , float(self.lineas[-1].alimento.costo) + self.lineas[-1].costo_extendido))
        subtotal = 0
        for linea in self.lineas:
            subtotal += (float(cantidad)*(float(linea.alimento.costo) + linea.costo_extendido))
        self.string_subtotal.set("{:8}:  {:4}".format("Subtotal","%0.2f"%subtotal))
        
        self.string_total.set("{:8}:  {:4}".format("Total","%0.2f"%(subtotal*1.16)))
        self.string_iva.set("{:8}:  {:4}".format("IVA","%0.2f"%((subtotal*1.16)-subtotal)))
        self.focusFamilia()
        print(self.lineas)


    def focusFamilia(self , *args): self.familia.focus()
    def focusNumero(self , *args): self.numero.focus()
    def focusUsuario(self , *args): self.usuario.focus()
        
    def desplegarArticulo(self , *args):
        usuario = self.usuario.get()
        familia = self.familia.get()
        numero  = self.numero.get()
        print(usuario , familia , numero)
        #self.familia.delete(0,END)
        #self.numero.delete(0,END)
        

        
        if usuario and familia and numero:
            ent_sku = int(int(familia)*1000 + int(numero))
      
            try:
                alimento = session.query(Alimento).filter(Alimento.sku == ent_sku).one()
            except:
                alimento = 0
            self.usuario.config(state = DISABLED)    
            if not alimento:
                mbox.showwarning("Busqueda", "El articulo no existe .")

                self.focusFamilia()
            
            else:
                nombre = str(alimento.nombre)
       
                
                #print(self.lineas[-1].alimento.nombre)
                precio = float(alimento.costo) *1.3
                #print(self.redondear_operacion(precio, base = 0.1))
        
                self.string_articulo.set("{} -- ${:3}".format(nombre , "%0.2f"%precio))

                self.cantidad.delete(0,END)
                self.cantidad.config(state = NORMAL)
                self.cantidad.focus()
                print(alimento.nombre)
        else :
            mbox.showwarning("Campos vacios", "Favor de llenar campos Usuario , Familia , y Numero .")
            if not usuario:
                self.focusUsuario()
            else:
                self.focusFamilia()




        print("{}-{}".format(familia , numero))
    def onSelectArticulos(self , *args):
        self.numero.focus()

                     
    def Alimento(self):
        i = 0
        j = 0
        self.destruir()
        lbl_busqueda = Label(self , text = "Buscar Alimentos " , font = (None , 12)) 
        lbl_busqueda.grid(row = 1 , column = 0)
        self.busqueda = Entry(self)
        self.busqueda.bind('<Return>', self.buscarAlimento)
        self.busqueda.grid(row = 1 , column = 1)

        
        self.alimentos_lb = Listbox(self)
        self.alimentos_lb.bind("<<ListboxSelect>>", self.onSelectAlimento)
        self.update_alimento_lb()
            


        self.alimentos_lb.grid(row = 2 , column = 0 ,columnspan = 2, rowspan = 6 , padx = 0 , sticky = 'NSWE')

        lbl_titulo = Label(self , text = "Alta de Alimento " , font = (None , 30))
        lbl_titulo.grid(row = 0 , column = 3 , columnspan = 3)
        
        lbl_nombre = Label(self , text = "Nombre " , font = (None , 12))
        lbl_nombre.grid(row = 1 , column = 2)
        
        self.nombre = Entry(self)
        self.nombre.grid(row = 1 , column =3 , columnspan = 3 , sticky = 'WE')

        lbl_sku = Label(self , text = "Familia " , font = (None , 12))
        lbl_sku.grid(row = 2 , column = 2)
        
        self.sku = Entry(self)
        self.sku.grid(row = 2 , column =3 , columnspan = 3 , sticky = 'WE')

        lbl_cantidad = Label(self , text = "Cantidad " , font = (None , 12))
        lbl_cantidad.grid(row = 3 , column = 2)
        
        self.cantidad = Entry(self)
        self.cantidad.grid(row = 3 , column =3 , columnspan = 3 , sticky = 'WE')
                
        lbl_costo = Label(self , text = "Costo " , font = (None , 12))
        lbl_costo.grid(row = 4 , column = 2)
        
        self.costo = Entry(self)
        self.costo.grid(row = 4 , column =3 , columnspan = 3 , sticky = 'WE')
        
        alta = Button(self , text = "Dar de alta" , command = self.boton_alta_alimento)
        alta.grid(row = 5 , column = 3 , sticky = 'NEW')

        actualizar = Button(self , text = "Actualizar" , command = self.actualizar_alimento)
        actualizar.grid(row = 5 , column = 5 , sticky = 'NEW')

        baja = Button(self , text = 'Eliminar' , command = self.eliminar_alimento)
        baja.grid(row = 5 , column = 4 , sticky = 'NEW')
        
        self.widgets_alimento = [lbl_busqueda , self.busqueda , self.alimentos_lb ,
                                 lbl_titulo , lbl_nombre , self.nombre ,
                                 lbl_sku , self.sku , lbl_cantidad , self.cantidad ,
                                 lbl_costo , self.costo , alta , actualizar , baja ,]
    def update_alimento_lb(self):
        alimentos = session.query(Alimento).all()
        self.alimentos_lb.delete(0,END)
        for alimento in alimentos:
            self.alimentos_lb.insert(END,"{:6} - {}".format(alimento.sku , alimento.nombre))
        for i in range(0,self.alimentos_lb.size(),2):
            self.alimentos_lb.itemconfigure(i,background = '#f0f0ff')
    def buscarAlimento(self , event):
        clave = self.busqueda.get()
        flag = 1
        try:
            int(clave)
        except:
            flag = 0
        if not flag:

            alimentos = session.query(Alimento).filter(Alimento.nombre.like('%'+clave+'%'))
            print(alimentos)
            if not alimentos.count(): mbox.showwarning("Busqueda", "La busqueda no produjo resultados .")
            self.alimentos_lb.delete(0,self.alimentos_lb.size()+1)
            for alimento in alimentos:
                self.alimentos_lb.insert(END,"{:6} - {}".format(alimento.sku , alimento.nombre))
            for i in range(0,self.alimentos_lb.size(),2):
                self.alimentos_lb.itemconfigure(i,background = '#f0f0ff')
        else :
            familia = session.query(Alimento).filter(Alimento.sku.like('%'+str(clave)+'%'))
            if not familia.count(): mbox.showwarning("Busqueda", "La busqueda no produjo resultados .")
            self.alimentos_lb.delete(0,END)
            for alimento in familia:
                self.alimentos_lb.insert(END,"{:6} - {}".format(alimento.sku , alimento.nombre))
            for i in range(0,self.alimentos_lb.size(),2):
                self.alimentos_lb.itemconfigure(i,background = '#f0f0ff')



    def onSelectAlimento(self , val):
        sender = val.widget
        idx = sender.curselection()
        nombre = sender.get(idx)
        print(nombre[9:])
        alimento = session.query(Alimento).filter(Alimento.nombre == nombre[9:]).one()

        self.nombre.delete(0,END)
        self.nombre.insert(0,alimento.nombre)

        self.sku.delete(0,END)
        self.sku.insert(0,int(int(alimento.sku)/1000))

        self.cantidad.delete(0,END)
        self.cantidad.insert(0,alimento.cantidad)

        self.costo.delete(0,END)
        self.costo.insert(0,alimento.costo)
    def familias(self):
        dict1 = {}

        with open('config.txt','r') as archivo:
            for linea in archivo:
                dict1[linea.split('\t')[0]] = linea.split('\t')[1][:-1]    

        return dict1
    def boton_alta_alimento(self):
        nombre = self.nombre.get().lower()
        sku = self.sku.get()
        cantidad = self.cantidad.get()
        costo = self.costo.get()
        igual = 0
        



        if str(sku) in self.diccionario.keys():

            try:
                nombre_igual = session.query(Alimento).filter(Alimento.nombre == nombre).one()
                print("nombre alimento = %s"%nombre_igual.nombre)
            except:
                igual = 1
        
            if not igual: mbox.showerror("Error", "El alimento ya existe ! \n favor de ingresar otro nombre")

            if nombre and sku and cantidad and costo and igual:
                try :
                    int(sku)
                except:
                    sku = 0

                if not sku:
                	mbox.showerror("Error", "El numero de familia no es numerico")
                else:
                    print("SKU = %s"%sku)
                    max_sku = session.query(func.max(Alimento.sku)).filter(Alimento.sku/1000 == int(sku)).one()
                
                    print("max_sku %s"%max_sku)
                    if max_sku[0] == None : max_sku = 0
                    if not max_sku:
                        sku_act = int("%s%s"%(sku,"001"))
                    else:
                        sku_act = int(int(max_sku[0])+1)     
                
                    alimento = Alimento(nombre = nombre , sku = sku_act, cantidad = cantidad , costo = costo)
                    session.add(alimento)
                    session.commit()
                    mbox.showinfo("Alta Exitosa", "Alimento %s dado de alta!"%alimento.nombre)     
                    self.update_alimento_lb()
            else :
                faltantes = []
                if not nombre:
                    faltantes.append("nombre ")
                if not sku:
                    faltantes.append("sku ")
                if not cantidad:
                    faltantes.append("cantidad ")
                if not costo:
                    faltantes.append("costo ")
                if not igual:
                    faltantes.append("\n Y el Nombre es igual")

                string = ''
                for faltante in faltantes:
                    string += faltante        


                mbox.showerror("Error", "Porfavor llenar %s"%string)
        else :
            mbox.showerror("Error", "Familia no existe")
                    

    def actualizar_alimento(self):
        nombre = self.nombre.get()
        sku = self.sku.get()
        cantidad = self.cantidad.get()
        costo = self.costo.get()
        
        if nombre and sku and cantidad and costo:
            alimento = session.query(Alimento).filter(Alimento.nombre == nombre).one()
            alimento.nombre = nombre
            alimento.sku = sku
            alimento.cantidad = cantidad
            alimento.costo = costo
            session.add(alimento)
            session.commit()
            mbox.showinfo("Actualizacion Exitosa", "Alimento %s Modificado!"%alimento.nombre)     
        else :
            faltantes = []
            if not nombre:
                faltantes.append("nombre ")
            if not sku:
                faltantes.append("sku ")
            if not cantidad:
                faltantes.append("cantidad ")
            if not costo:
                faltantes.append("costo ")

            string = ''
            for faltante in faltantes:
                string += faltante        


            mbox.showerror("Error", "Porfavor llenar %s"%string)

    def eliminar_alimento(self):
        nombre = self.nombre.get()
        alimento = session.query(Alimento).filter(Alimento.nombre == nombre).one()
        #session.delete(usuario)
        a = mbox.askquestion("Eliminar Alimento", "Estas seguro que quieres borrar el alimento %s ?"%alimento.nombre)
        print('respuesta = %s'%a)
        if a=='yes':
            session.delete(alimento)
            session.commit()  
        self.update_alimento_lb()            





    def onExit(self,*args):
        self.quit()
        
       
        



    def onSelectUser(self , val):
        sender = val.widget
        idx = sender.curselection()
        nombre = sender.get(idx)
        usuario = session.query(Usuario).filter(Usuario.nombre == nombre).one()
        print(usuario.email)
        self.nombre.delete(0,END)
        self.nombre.insert(0,usuario.nombre)

        self.direccion.delete(0,END)
        self.direccion.insert(0,usuario.direccion)

        self.email.delete(0,END)
        self.email.insert(0,usuario.email)

        self.telefono.delete(0,END)
        self.telefono.insert(0,usuario.telefono)

        

    def buscarUsuarios(self,event ):
        clave = self.busqueda.get()
        
        usuarios = session.query(Usuario).filter(Usuario.nombre.like('%'+clave+'%'))
        print(usuarios)
        if not usuarios.count(): mbox.showwarning("Busqueda", "La busqueda no produjo resultados .")
        self.usuarios_lb.delete(0,self.usuarios_lb.size()+1)
        for usuario in usuarios:
            self.usuarios_lb.insert(END,usuario.nombre)
        for i in range(0,self.usuarios_lb.size(),2):
            self.usuarios_lb.itemconfigure(i,background = '#f0f0ff')    
        
        
    def actualizar_usuario(self):
        nombre = self.nombre.get()
        direccion = self.direccion.get()
        email = self.email.get()
        telefono = self.telefono.get()
        print(nombre , direccion , email , telefono)
        if nombre and direccion and email and telefono:
            usuario = session.query(Usuario).filter(Usuario.nombre == nombre).one()
            usuario.nombre = nombre
            usuario.direccion = direccion
            usuario.email = email
            usuario.telefono = telefono
            session.add(usuario)
            session.commit()
            mbox.showinfo("Actualizacion Exitosa", "Usuario %s Modificado!"%usuario.nombre)     
        else :
            faltantes = []
            if not nombre:
                faltantes.append("nombre ")
            if not direccion:
                faltantes.append("direccion ")
            if not email:
                faltantes.append("email ")
            if not telefono:
                faltantes.append("telefono ")

            string = ''
            for faltante in faltantes:
                string += faltante        


            mbox.showerror("Error", "Porfavor llenar %s"%string)

    def update_usuarios_lb(self):
        self.usuarios_lb.delete(0,END)
        for usuario in session.query(Usuario):
            self.usuarios_lb.insert(END,usuario.nombre)

        for i in range(0,self.usuarios_lb.size(),2):
            self.usuarios_lb.itemconfigure(i,background = '#f0f0ff')            
    def alta_usuario(self):
        self.destruir()

        

        lbl_busqueda = Label(self , text = "Buscar Usuarios " , font = (None , 12)) 
        lbl_busqueda.grid(row = 1 , column = 0)
        self.busqueda = Entry(self)
        self.busqueda.bind('<Return>', self.buscarUsuarios)
        self.busqueda.grid(row = 1 , column = 1)
        lbl_busqueda.bind("<Control-c>",self.onExit)
        
        self.usuarios_lb = Listbox(self)
        self.usuarios_lb.bind("<<ListboxSelect>>", self.onSelectUser)
        self.update_usuarios_lb()
            


        self.usuarios_lb.grid(row = 3 , column = 1 , rowspan = 6 , padx = 50)

        lbl_titulo = Label(self , text = "Alta de Usuario " , font = (None , 30))
        lbl_titulo.grid(row = 0 , column = 3 , columnspan = 3)
        
        lbl_nombre = Label(self , text = "Nombre " , font = (None , 12))
        lbl_nombre.grid(row = 1 , column = 2)
        
        self.nombre = Entry(self)
        self.nombre.grid(row = 1 , column =3 , columnspan = 3 , sticky = 'WE')

        lbl_direccion = Label(self , text = "Direccion " , font = (None , 12))
        lbl_direccion.grid(row = 2 , column = 2)
        
        self.direccion = Entry(self)
        self.direccion.grid(row = 2 , column =3 , columnspan = 3 , sticky = 'WE')

        lbl_email = Label(self , text = "Email " , font = (None , 12))
        lbl_email.grid(row = 3 , column = 2)
        
        self.email = Entry(self)
        self.email.grid(row = 3 , column =3 , columnspan = 3 , sticky = 'WE')
                
        lbl_telefono = Label(self , text = "Telefono " , font = (None , 12))
        lbl_telefono.grid(row = 4 , column = 2)
        
        self.telefono = Entry(self)
        self.telefono.grid(row = 4 , column =3 , columnspan = 3 , sticky = 'WE')
        
        alta = Button(self , text = "Dar de alta" , command = self.boton_alta_usuario)
        alta.grid(row = 5 , column = 3 , sticky = 'NEW')

        actualizar = Button(self , text = "Actualizar" , command = self.actualizar_usuario)
        actualizar.grid(row = 5 , column = 5 , sticky = 'NEW')

        baja = Button(self , text = 'Eliminar' , command = self.eliminar_usuario)
        baja.grid(row = 5 , column = 4 , sticky = 'NEW')

        #baja2 = Button(self , text = 'Eliminarddd' , command = self.destruir)
        #baja2.grid(row = 10 , column = 4 , sticky = 'NEW')
        
        self.widgets_alta_usuario = [lbl_titulo , lbl_nombre , self.nombre , 
                                     lbl_direccion , self.direccion ,
                                     lbl_email , self.email , 
                                     lbl_telefono , self.telefono , 
                                     alta ,lbl_busqueda , self.busqueda , 
                                     self.usuarios_lb ,baja , actualizar , ]
        
        #total_usuarios = session.query(func.count(Usuario.nombre)).one()
    def eliminar_usuario(self):
        nombre = self.nombre.get()
        usuario = session.query(Usuario).filter(Usuario.nombre == nombre).one()
        #session.delete(usuario)
        a = mbox.askquestion("Eliminar Usuario", "Estas seguro que quieres borrar al usuario %s ?"%usuario.nombre)
        print('respuesta = %s'%a)
        if a=='yes':
            session.delete(usuario)
            session.commit()
        self.update_usuarios_lb()    

        
    def boton_alta_usuario(self):
        #global nombre , direccion , email , telefono
        nombre = self.nombre.get().lower()
        direccion = self.direccion.get()
        email = self.email.get()
        telefono = self.telefono.get()
        print(nombre , direccion , email , telefono)
        if nombre and direccion and email and telefono:
            usuario = Usuario(nombre = nombre , direccion = direccion , email = email , telefono = telefono)
            session.add(usuario)
            session.commit()
            mbox.showinfo("Alta Exitosa", "Usuario %s dado de alta!"%usuario.nombre)     
            self.update_usuarios_lb()
        else :
            faltantes = []
            if not nombre:
                faltantes.append("nombre ")
            if not direccion:
                faltantes.append("direccion ")
            if not email:
                faltantes.append("email ")
            if not telefono:
                faltantes.append("telefono ")

            string = ''
            for faltante in faltantes:
                string += faltante        


            mbox.showerror("Error", "Porfavor llenar %s"%string)
        
       
        

    def destruir(self):

        if self.widgets_alta_usuario:
            for i in self.widgets_alta_usuario:
                i.destroy()        
            self.widgets_alta_usuario = []
        if self.widgets_alimento:
            for i in self.widgets_alimento:
                i.destroy()        
            self.widgets_alimento = []  

        if self.widgets_orden:
            for i in self.widgets_orden:
                i.destroy()        
            self.widgets_orden = [] 
        if self.widgets_ordenes_no_enviadas:
            for i in self.widgets_ordenes_no_enviadas:
                i.destroy()
            self.widgets_ordenes_no_enviadas = []  
        if self.widgets_grafica:
            for i in self.widgets_grafica:
                i.destroy()
            for i in  self.checkbuttons:
                i.destroy()    
            self.widgets_grafica = []    
        if self.widgets_estadistica:
            for i in self.widgets_estadistica:
                i.destroy()
            for i in self.botones:
                i.destroy()    
            self.widgets_estadistica = []          

def main():
    root = Tk()
    app = Aplicacion(root)
    root.bind("<Control-q>", app.onExit)
    root.mainloop()


if __name__ == '__main__':
    main() 
